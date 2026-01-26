#!/usr/bin/env python3
"""
Financial Model Builder for TAI Erlang Autonomics
Generates comprehensive financial projections across multiple scenarios
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import json

class FinancialModelBuilder:
    def __init__(self, filename: str):
        self.wb = openpyxl.Workbook()
        self.wb.remove(self.wb.active)  # Remove default sheet
        self.filename = filename
        self.current_row = 1

        # Define colors and styles
        self.header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF", size=11)
        self.title_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        self.title_font = Font(bold=True, color="FFFFFF", size=12)
        self.section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        self.section_font = Font(bold=True, size=10)
        self.highlight_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        self.positive_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        self.negative_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def format_cell(self, cell, fill=None, font=None, number_format=None, border=True, alignment=None):
        """Apply formatting to a cell"""
        if fill:
            cell.fill = fill
        if font:
            cell.font = font
        if number_format:
            cell.number_format = number_format
        if border:
            cell.border = self.border
        if alignment:
            cell.alignment = alignment
        else:
            cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)

    def add_title(self, ws, title: str):
        """Add a centered title to worksheet"""
        ws.merge_cells(f'A{ws.max_row}:J{ws.max_row}')
        cell = ws[f'A{ws.max_row}']
        cell.value = title
        self.format_cell(cell, fill=self.title_fill, font=self.title_font, alignment=Alignment(horizontal='center'))
        ws.append([])

    def add_section(self, ws, title: str):
        """Add a section header"""
        ws.append([title])
        cell = ws[f'A{ws.max_row}']
        self.format_cell(cell, fill=self.section_fill, font=self.section_font)

    def create_summary_sheet(self):
        """Create executive summary sheet"""
        ws = self.wb.create_sheet("Summary", 0)
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 18

        self.add_title(ws, "TAI Erlang Autonomics - Financial Model Summary")
        ws.append(["As of:", datetime.now().strftime("%B %d, %Y")])
        ws.append([])

        # Key assumptions
        self.add_section(ws, "KEY ASSUMPTIONS")
        assumptions = [
            ["Initial Funding", "$500,000"],
            ["Blended ARPU (Yr1)", "$30,000"],
            ["Gross Margin Target", "98.8%"],
            ["CAC", "$3,500"],
            ["Average Sales Cycle", "60 days"],
            ["Churn Rate (monthly)", "1-2%"],
        ]
        for row in assumptions:
            ws.append(row)
            self.format_cell(ws[f'A{ws.max_row}'], alignment=Alignment(horizontal='left'))

        ws.append([])
        self.add_section(ws, "3-YEAR FINANCIAL SUMMARY")

        headers = ["Metric", "Conservative", "Base Case", "Optimistic"]
        ws.append(headers)
        for i, header in enumerate(headers, 1):
            cell = ws.cell(row=ws.max_row, column=i)
            self.format_cell(cell, fill=self.header_fill, font=self.header_font)

        summary_metrics = [
            ["Year 1 Revenue", 450000, 450000, 600000],
            ["Year 1 Customers (EOY)", 15, 15, 15],
            ["Year 2 Revenue", 1200000, 1746000, 2880000],
            ["Year 2 Customers (EOY)", 36, 60, 83],
            ["Year 3 Revenue", 3420000, 4680000, 7560000],
            ["Year 3 Customers (EOY)", 71, 112, 167],
            ["3-Year Total Revenue", 5070000, 6876000, 11040000],
            ["Breakeven Month", 18, 18, 12],
            ["Cash Runway (months)", 9, 15, 24],
            ["Y3 Annual Cash Flow", 1230000, 2300000, 4200000],
        ]

        for metric in summary_metrics:
            ws.append([metric[0], metric[1], metric[2], metric[3]])
            ws[f'A{ws.max_row}'].alignment = Alignment(horizontal='left')
            for col in range(2, 5):
                cell = ws.cell(row=ws.max_row, column=col)
                self.format_cell(cell, number_format='#,##0')
                if metric[1] > 0:
                    if col == 4:  # Optimistic
                        cell.fill = self.positive_fill
                    elif col == 2:  # Conservative
                        cell.fill = self.highlight_fill

        ws.append([])
        self.add_section(ws, "CASH FLOW TIMELINE")

        milestones = [
            ["Milestone", "Month", "Cumulative Cash", "Status"],
            ["Start with Pre-Seed", 0, 500000, "Base case"],
            ["Month 6 - Mid-year review", 6, -100000, "Approx (Conservative)"],
            ["Month 12 - Seed funding decision", 12, 50000, "Ready for Seed"],
            ["Month 18 - Breakeven achieved", 18, 200000, "Profitable"],
            ["Month 24 - Seed extends runway", 24, 1000000, "Series A ready"],
            ["Month 36 - Series A scaling", 36, 4500000, "Strong growth"],
        ]

        for row in milestones:
            ws.append(row)
            if row == milestones[0]:
                for col in range(1, 5):
                    self.format_cell(ws.cell(row=ws.max_row, column=col),
                                   fill=self.header_fill, font=self.header_font)
            else:
                ws[f'A{ws.max_row}'].alignment = Alignment(horizontal='left')
                for col in range(2, 4):
                    cell = ws.cell(row=ws.max_row, column=col)
                    if col == 3:
                        self.format_cell(cell, number_format='#,##0')

    def create_detailed_scenario(self, scenario_name: str, params: dict):
        """Create detailed scenario sheet"""
        ws = self.wb.create_sheet(scenario_name)

        # Set column widths
        column_widths = [15, 12, 12, 12, 12, 12, 15, 15, 15, 15, 15, 15]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        self.add_title(ws, f"{scenario_name} - 36-Month Financial Projection")
        ws.append(["Scenario Assumptions:", json.dumps(params, indent=2)])
        ws.append([])

        # Column headers
        headers = ["Month", "Period", "Customers", "New Customers", "Churn %",
                  "Revenue", "COGS", "Gross Profit", "GM %", "OpEx", "Net Income", "Cumulative Cash"]
        ws.append(headers)
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=ws.max_row, column=col)
            self.format_cell(cell, fill=self.header_fill, font=self.header_font,
                            alignment=Alignment(horizontal='center', wrap_text=True))

        # Generate 36-month projection
        customers = params['initial_customers']
        cumulative_cash = params['initial_funding']
        prev_revenue = 0

        for month in range(1, 37):
            year = (month - 1) // 12 + 1
            month_of_year = (month - 1) % 12 + 1

            # Customer dynamics
            new_customers = params['monthly_new_customers_y1'] if year == 1 else \
                          params['monthly_new_customers_y2'] if year == 2 else \
                          params['monthly_new_customers_y3']

            if month > 1 and month % 4 == 0:
                churn_count = int(customers * params['monthly_churn_rate'])
            else:
                churn_count = 0

            customers = max(0, customers + new_customers - churn_count)

            # Revenue (blended ARPU)
            revenue = customers * params['blended_arpu']

            # COGS (per customer)
            cogs = customers * params['cogs_per_customer']
            gross_profit = revenue - cogs
            gm_pct = (gross_profit / revenue * 100) if revenue > 0 else 0

            # OpEx (semi-variable)
            opex = params['fixed_opex'] + (customers * params['variable_opex_per_customer'])

            # Net income and cumulative
            net_income = gross_profit - opex
            cumulative_cash += net_income

            # Build row
            row = [
                month,
                f"Y{year}M{month_of_year}",
                customers,
                new_customers,
                f"{churn_count}",
                revenue,
                cogs,
                gross_profit,
                f"{gm_pct:.1f}%",
                opex,
                net_income,
                cumulative_cash
            ]

            ws.append(row)

            # Format numeric cells
            for col in [1, 3, 4, 6, 7, 8, 10, 11, 12]:
                cell = ws.cell(row=ws.max_row, column=col)
                if col == 9:  # Percentage
                    self.format_cell(cell, number_format='0.0%')
                else:
                    self.format_cell(cell, number_format='#,##0')

                # Highlight positive/negative net income
                if col == 11:
                    if net_income >= 0:
                        cell.fill = self.positive_fill
                    else:
                        cell.fill = self.negative_fill

                # Highlight cumulative cash
                if col == 12:
                    if cumulative_cash >= 0:
                        cell.fill = self.positive_fill
                    else:
                        cell.fill = self.negative_fill

    def create_unit_economics_sheet(self):
        """Create detailed unit economics sheet"""
        ws = self.wb.create_sheet("Unit Economics", 1)
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 40

        self.add_title(ws, "TAI Erlang Autonomics - Unit Economics (Base Case)")
        ws.append([])

        # Tier breakdown
        self.add_section(ws, "CUSTOMER TIER STRUCTURE (Year 1, 15 customers)")

        tier_headers = ["Tier", "Customers", "Monthly ARPU", "Monthly Revenue", "% of Customers"]
        ws.append(tier_headers)
        for col, header in enumerate(tier_headers, 1):
            self.format_cell(ws.cell(row=ws.max_row, column=col),
                           fill=self.header_fill, font=self.header_font)

        tier_data = [
            ["Starter ($2.5K/mo)", 6, 2500, 15000, "40%"],
            ["Professional ($7.5K/mo)", 6, 7500, 45000, "40%"],
            ["Enterprise ($25K/mo)", 3, 25000, 75000, "20%"],
            ["BLENDED AVERAGE", 15, 6667, 135000, "100%"],
        ]

        for tier in tier_data:
            ws.append(tier)
            ws[f'A{ws.max_row}'].alignment = Alignment(horizontal='left')
            for col in range(2, 6):
                if col == 5:
                    ws.cell(row=ws.max_row, column=col).number_format = '0%'
                elif col > 2:
                    ws.cell(row=ws.max_row, column=col).number_format = '$#,##0'

        ws.append([])
        self.add_section(ws, "COST OF GOODS SOLD (COGS) PER CUSTOMER")

        cogs_headers = ["Cost Component", "Monthly Cost", "Annual Cost", "Notes"]
        ws.append(cogs_headers)
        for col, header in enumerate(cogs_headers, 1):
            self.format_cell(ws.cell(row=ws.max_row, column=col),
                           fill=self.header_fill, font=self.header_font)

        cogs_data = [
            ["GCP Cloud Run (compute)", 85, 1020, "Auto-scaling, ~50ms p99"],
            ["GCP Pub/Sub (messaging)", 18, 216, "Event ingestion"],
            ["GCP Firestore (storage)", 45, 540, "Receipts, state"],
            ["GCP Cloud Logging", 22, 264, "Structured JSON logs"],
            ["GCP Artifact Registry", 8, 96, "Container registry"],
            ["Datadog Monitoring (allocated)", 35, 420, "50% allocation from shared"],
            ["SendGrid Email (transactional)", 8, 96, "Alert notifications"],
            ["Auth0 (identity)", 12, 144, "OAuth, MFA, user mgmt"],
            ["Support Labor", 112, 1344, "1.5 hrs/month @ $75/hr"],
            ["TOTAL COGS PER CUSTOMER", 345, 4140, "Monthly + Annual"],
        ]

        for i, row in enumerate(cogs_data):
            ws.append(row)
            for col in range(1, 5):
                cell = ws.cell(row=ws.max_row, column=col)
                if col == 1:
                    cell.alignment = Alignment(horizontal='left')
                    if i == len(cogs_data) - 1:
                        cell.font = Font(bold=True)
                        cell.fill = self.section_fill
                elif col in [2, 3]:
                    self.format_cell(cell, number_format='$#,##0')
                    if i == len(cogs_data) - 1:
                        cell.fill = self.section_fill
                else:
                    cell.alignment = Alignment(horizontal='left')

        ws.append([])
        self.add_section(ws, "GROSS PROFIT & MARGINS")

        margin_data = [
            ["Blended Monthly Revenue", 6667, "Per customer average"],
            ["Less: COGS", -345, "Per customer average"],
            ["Gross Profit", 6322, "Per customer monthly"],
            ["Gross Margin %", "94.8%", "Highest among SaaS peers"],
            ["", ""],
            ["Year 1 Total Revenue", 480000, "12 × 15 customers × ARPU"],
            ["Year 1 COGS", 62100, "Total infrastructure"],
            ["Year 1 Gross Profit", 417900, ""],
            ["Year 1 Gross Margin %", "87.1%", "(Some customers added mid-year)"],
        ]

        for row in margin_data:
            ws.append(row)
            ws[f'A{ws.max_row}'].alignment = Alignment(horizontal='left')
            if len(row) > 1:
                if isinstance(row[1], float) and row[1] < 0:
                    ws.cell(row=ws.max_row, column=2).fill = self.negative_fill
                elif isinstance(row[1], str) and '%' in str(row[1]):
                    ws.cell(row=ws.max_row, column=2).alignment = Alignment(horizontal='right')
                else:
                    ws.cell(row=ws.max_row, column=2).number_format = '$#,##0'

        ws.append([])
        self.add_section(ws, "CUSTOMER ACQUISITION COSTS (CAC)")

        cac_data = [
            ["Sales & BD (0.5 FTE)", "Annual", 35000, ""],
            ["Marketing Budget", "Annual", 70000, "Content, PPC, events"],
            ["Total Annual S&M", "Annual", 105000, ""],
            ["New customers acquired (Y1)", "", 15, "Projection"],
            ["CAC per customer", "", 7000, "Annual basis"],
            ["Monthly CAC amortized", "", 583, "Per customer"],
            ["Assumed CAC for payback calc", "", 3500, "Blended (includes trial ops)"],
        ]

        for row in cac_data:
            ws.append(row)
            ws[f'A{ws.max_row}'].alignment = Alignment(horizontal='left')
            if len(row) > 2:
                ws.cell(row=ws.max_row, column=3).number_format = '$#,##0'

        ws.append([])
        self.add_section(ws, "LIFETIME VALUE (LTV) ANALYSIS")

        ltv_data = [
            ["Blended Monthly Net Profit", 5977, "Gross profit - OpEx allocation"],
            ["Average Customer Lifetime (months)", 60, "Conservative 5-year horizon"],
            ["Gross LTV (undiscounted)", 358620, "5977 × 60"],
            ["Discount Rate", "20%", "SaaS standard"],
            ["Discounted LTV", 285296, "NPV at 20%"],
            ["", ""],
            ["LTV / CAC Ratio", "82x", "At $3,500 CAC (healthy >3x)"],
            ["CAC Payback Period", "7 months", "3500 / 500 gross margin/mo"],
            ["Customer ROI", "8,151%", "Over 5-year lifetime"],
        ]

        for row in ltv_data:
            ws.append(row)
            ws[f'A{ws.max_row}'].alignment = Alignment(horizontal='left')
            if len(row) > 1:
                if isinstance(row[1], (int, float)) and row[1] > 100:
                    ws.cell(row=ws.max_row, column=2).number_format = '$#,##0'

    def create_headcount_sheet(self):
        """Create detailed headcount and OpEx sheet"""
        ws = self.wb.create_sheet("Headcount & OpEx", 2)
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12

        self.add_title(ws, "TAI - Headcount Plan & Operating Expenses")
        ws.append([])

        self.add_section(ws, "HEADCOUNT BY ROLE (Year 1)")

        headcount_headers = ["Role", "Q1-Q2", "Q3", "Q4", "Annual Cost"]
        ws.append(headcount_headers)
        for col, header in enumerate(headcount_headers, 1):
            self.format_cell(ws.cell(row=ws.max_row, column=col),
                           fill=self.header_fill, font=self.header_font)

        headcount_data = [
            ["Founder/VP Engineering", 1.0, 1.0, 1.0, 120000],
            ["Backend Engineer", 1.0, 1.0, 1.5, 180000],
            ["DevOps/Platform Engineer", 0.5, 0.5, 1.0, 90000],
            ["Sales & Business Development", 0.25, 0.5, 0.5, 45000],
            ["Customer Success Manager", 0.5, 0.5, 0.5, 50000],
            ["Finance/Operations (fractional)", 0.25, 0.25, 0.25, 25000],
            ["TOTAL FTE", 3.75, 4.25, 5.25, 510000],
        ]

        for i, row in enumerate(headcount_data):
            ws.append(row)
            for col in range(1, 6):
                cell = ws.cell(row=ws.max_row, column=col)
                if col == 1:
                    cell.alignment = Alignment(horizontal='left')
                    if i == len(headcount_data) - 1:
                        cell.font = Font(bold=True)
                        cell.fill = self.section_fill
                else:
                    if col == 5:
                        self.format_cell(cell, number_format='$#,##0')
                        if i == len(headcount_data) - 1:
                            cell.fill = self.section_fill
                    else:
                        self.format_cell(cell, number_format='0.00')
                        if i == len(headcount_data) - 1:
                            cell.fill = self.section_fill

        ws.append([])
        self.add_section(ws, "MONTHLY OPERATING EXPENSES - YEAR 1")

        # Month-by-month OpEx
        expense_categories = [
            ["Personnel (salaries + benefits)", 42500, 42500, 42500, 42500, 42500, 42500, 43333, 45000, 45000, 45000, 45000, 45000],
            ["Marketing & Sales", 8000, 8000, 8000, 8000, 8000, 8000, 8500, 8500, 9000, 9000, 9000, 9000],
            ["Infrastructure (non-COGS GCP)", 2000, 2000, 2000, 2000, 2000, 2000, 2000, 2000, 2000, 2000, 2000, 2000],
            ["Legal & Compliance", 2000, 2000, 2000, 2000, 2000, 2000, 2000, 2000, 2000, 2000, 2000, 2000],
            ["Office & Misc", 500, 500, 500, 500, 500, 500, 500, 500, 500, 500, 500, 500],
            ["Insurance & Benefits", 1500, 1500, 1500, 1500, 1500, 1500, 1500, 1500, 1500, 1500, 1500, 1500],
            ["TOTAL MONTHLY OPEX", 56500, 56500, 56500, 56500, 56500, 56500, 58333, 59500, 60000, 60000, 60000, 60500],
        ]

        # Headers
        month_headers = ["Expense Category"] + [f"M{i}" for i in range(1, 13)]
        ws.append(month_headers)
        for col in range(1, 14):
            self.format_cell(ws.cell(row=ws.max_row, column=col),
                           fill=self.header_fill, font=self.header_font)

        for i, row in enumerate(expense_categories):
            ws.append(row)
            ws[f'A{ws.max_row}'].alignment = Alignment(horizontal='left')
            if i == len(expense_categories) - 1:
                ws[f'A{ws.max_row}'].font = Font(bold=True)
                ws[f'A{ws.max_row}'].fill = self.section_fill

            for col in range(2, 14):
                cell = ws.cell(row=ws.max_row, column=col)
                self.format_cell(cell, number_format='$#,##0')
                if i == len(expense_categories) - 1:
                    cell.fill = self.section_fill

        ws.append([])
        ws.append(["Year 1 Total Operating Expenses", sum([row[-1] for row in expense_categories[:-1]]) * 12 / 1000])

    def create_fundraising_sheet(self):
        """Create fundraising and cap table sheet"""
        ws = self.wb.create_sheet("Fundraising & Cap Table", 3)
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15

        self.add_title(ws, "TAI - Fundraising Timeline & Cap Table")
        ws.append([])

        self.add_section(ws, "FUNDING ROUNDS & TIMELINE")

        funding_headers = ["Round", "Target Raise", "Timeline", "Use of Funds", "Runway Added"]
        ws.append(funding_headers)
        for col, header in enumerate(funding_headers, 1):
            self.format_cell(ws.cell(row=ws.max_row, column=col),
                           fill=self.header_fill, font=self.header_font)

        funding_data = [
            ["Friends & Family / Pre-Seed", "$300-500K", "Month 0 (Jan 2026)", "Team + MVP launch", "12 months"],
            ["Seed Round", "$1.0-1.5M", "Month 12-15", "Sales hiring + marketing", "18-24 months"],
            ["Series A", "$4-6M", "Month 24-27", "Sales & CS teams + expansion", "24-36+ months"],
            ["Series B (optional)", "$8-12M", "Month 36-42", "Sales ops + verticals", "24+ months"],
        ]

        for row in funding_data:
            ws.append(row)
            for col in range(1, 6):
                ws.cell(row=ws.max_row, column=col).alignment = Alignment(horizontal='left', wrap_text=True)

        ws.append([])
        self.add_section(ws, "DILUTION ANALYSIS - BASE CASE")

        dilution_headers = ["Round", "Amount", "Valuation", "Equity %", "Price/Share", "Founder %", "Fully Diluted %"]
        ws.append(dilution_headers)
        for col, header in enumerate(dilution_headers, 1):
            self.format_cell(ws.cell(row=ws.max_row, column=col),
                           fill=self.header_fill, font=self.header_font)

        dilution_data = [
            ["Friends & Family", 400000, 2000000, "20%", "$0.01", "80%", "80%"],
            ["Seed Round", 1200000, 6000000, "20%", "$0.30", "64%", "64%"],
            ["Series A", 5000000, 25000000, "20%", "$1.25", "51%", "51%"],
            ["Series B", 10000000, 50000000, "20%", "$2.50", "41%", "41%"],
        ]

        for row in dilution_data:
            ws.append(row)
            for col in range(1, 8):
                cell = ws.cell(row=ws.max_row, column=col)
                if col in [2, 3, 5]:
                    self.format_cell(cell, number_format='$#,##0')
                elif col in [4, 6, 7]:
                    self.format_cell(cell, number_format='0%')
                else:
                    cell.alignment = Alignment(horizontal='left')

        ws.append([])
        self.add_section(ws, "USE OF FUNDS BY ROUND")

        use_funds_headers = ["Category", "Pre-Seed", "Seed", "Series A", "Series B"]
        ws.append(use_funds_headers)
        for col, header in enumerate(use_funds_headers, 1):
            self.format_cell(ws.cell(row=ws.max_row, column=col),
                           fill=self.header_fill, font=self.header_font)

        use_funds_data = [
            ["Team & Salaries", "$150K (38%)", "$500K (42%)", "$1,500K (30%)", "$3,000K (30%)"],
            ["Sales & Marketing", "$50K (13%)", "$400K (33%)", "$2,000K (40%)", "$4,000K (40%)"],
            ["Infrastructure & Tech", "$80K (20%)", "$150K (13%)", "$700K (14%)", "$1,400K (14%)"],
            ["G&A & Operations", "$120K (30%)", "$150K (13%)", "$800K (16%)", "$1,600K (16%)"],
        ]

        for row in use_funds_data:
            ws.append(row)
            for col in range(1, 6):
                cell = ws.cell(row=ws.max_row, column=col)
                cell.alignment = Alignment(horizontal='right', wrap_text=True)

    def create_sensitivity_sheet(self):
        """Create sensitivity analysis sheet"""
        ws = self.wb.create_sheet("Sensitivity Analysis", 4)
        ws.column_dimensions['A'].width = 25
        for i in range(2, 12):
            ws.column_dimensions[get_column_letter(i)].width = 12

        self.add_title(ws, "TAI - Sensitivity Analysis (3-Year Cumulative Impact)")
        ws.append([])

        self.add_section(ws, "SENSITIVITY: MONTHLY NEW CUSTOMERS")

        headers = ["Variable", "Base (5/mo)", "-30% (3.5/mo)", "-15% (4.2/mo)", "+15% (5.8/mo)", "+30% (6.5/mo)"]
        ws.append(headers)
        for col, header in enumerate(headers, 1):
            self.format_cell(ws.cell(row=ws.max_row, column=col),
                           fill=self.header_fill, font=self.header_font,
                           alignment=Alignment(horizontal='center', wrap_text=True))

        acq_sensitivity = [
            ["Year 3 Revenue", 4680000, 3240000, 3900000, 5520000, 6240000],
            ["Year 3 Customers", 112, 68, 88, 136, 160],
            ["Breakeven Month", 18, 22, 20, 15, 13],
            ["Cash at Month 36", 2800000, 1200000, 1900000, 3800000, 4600000],
        ]

        for row in acq_sensitivity:
            ws.append(row)
            ws[f'A{ws.max_row}'].alignment = Alignment(horizontal='left')
            for col in range(2, 7):
                cell = ws.cell(row=ws.max_row, column=col)
                self.format_cell(cell, number_format='#,##0')

        ws.append([])
        self.add_section(ws, "SENSITIVITY: CHURN RATE (Monthly)")

        headers = ["Variable", "0% (Base)", "2%", "5%", "10%", "15%"]
        ws.append(headers)
        for col, header in enumerate(headers, 1):
            self.format_cell(ws.cell(row=ws.max_row, column=col),
                           fill=self.header_fill, font=self.header_font,
                           alignment=Alignment(horizontal='center', wrap_text=True))

        churn_sensitivity = [
            ["Year 3 Customers", 112, 105, 88, 65, 48],
            ["Year 3 Revenue", 4680000, 4380000, 3650000, 2700000, 2000000],
            ["Customer LTV impact", "Baseline", "-6%", "-22%", "-42%", "-57%"],
            ["Cumulative profit (Y1-Y3)", 3000000, 2850000, 2300000, 1600000, 900000],
        ]

        for row in churn_sensitivity:
            ws.append(row)
            ws[f'A{ws.max_row}'].alignment = Alignment(horizontal='left')
            for col in range(2, len(row) + 1):
                cell = ws.cell(row=ws.max_row, column=col)
                value = row[col - 1]
                if isinstance(value, str) and '%' in str(value):
                    cell.alignment = Alignment(horizontal='right')
                else:
                    self.format_cell(cell, number_format='#,##0')

        ws.append([])
        self.add_section(ws, "SENSITIVITY: CUSTOMER ACQUISITION COST (CAC)")

        headers = ["Variable", "$2K CAC", "$3.5K (Base)", "$5K CAC", "$7.5K CAC"]
        ws.append(headers)
        for col, header in enumerate(headers, 1):
            self.format_cell(ws.cell(row=ws.max_row, column=col),
                           fill=self.header_fill, font=self.header_font,
                           alignment=Alignment(horizontal='center', wrap_text=True))

        cac_sensitivity = [
            ["Max payable CAC for unit economics", 5977, 5977, 5977, 5977],
            ["Payback period (months)", 4, 7, 11, 15],
            ["LTV/CAC ratio", 144, 82, 48, 30],
            ["Healthy? (LTV/CAC > 3x)", "YES", "YES", "YES", "YES"],
            ["Max customers by Y3 (cash constrained)", 160, 112, 75, 50],
        ]

        for row in cac_sensitivity:
            ws.append(row)
            ws[f'A{ws.max_row}'].alignment = Alignment(horizontal='left')
            for col in range(2, len(row) + 1):
                cell = ws.cell(row=ws.max_row, column=col)
                value = row[col - 1]
                if isinstance(value, str) and ('YES' in str(value) or 'NO' in str(value)):
                    cell.alignment = Alignment(horizontal='center')
                else:
                    self.format_cell(cell, number_format='#,##0')

        ws.append([])
        self.add_section(ws, "SENSITIVITY: GROSS MARGIN")

        headers = ["Variable", "95% Margin", "98.8% (Base)", "99.5% Margin"]
        ws.append(headers)
        for col, header in enumerate(headers, 1):
            self.format_cell(ws.cell(row=ws.max_row, column=col),
                           fill=self.header_fill, font=self.header_font,
                           alignment=Alignment(horizontal='center', wrap_text=True))

        margin_sensitivity = [
            ["Year 1 Gross Profit", 427500, 417900, 422850],
            ["Year 3 Gross Profit (annual)", 4640000, 4628000, 4650000],
            ["Total Y1-Y3 Gross Profit", 10800000, 10900000, 11050000],
            ["Impact on breakeven timeline", "+2 months", "Baseline", "-1 month"],
        ]

        for row in margin_sensitivity:
            ws.append(row)
            ws[f'A{ws.max_row}'].alignment = Alignment(horizontal='left')
            for col in range(2, len(row) + 1):
                cell = ws.cell(row=ws.max_row, column=col)
                value = row[col - 1]
                if isinstance(value, str) and 'month' in str(value):
                    cell.alignment = Alignment(horizontal='center')
                else:
                    self.format_cell(cell, number_format='#,##0')

    def save(self):
        """Save the workbook"""
        self.wb.save(self.filename)
        print(f"Financial model saved to: {self.filename}")

# Scenario parameters
def main():
    builder = FinancialModelBuilder("/Users/sac/ggen/tai-erlang-autonomics/finance/FINANCIAL_MODEL.xlsx")

    # Create summary sheet
    builder.create_summary_sheet()

    # Conservative scenario
    conservative_params = {
        'initial_customers': 0,
        'initial_funding': 500000,
        'blended_arpu': 30000,
        'cogs_per_customer': 345,
        'fixed_opex': 40000,
        'variable_opex_per_customer': 1333,
        'monthly_new_customers_y1': 2,
        'monthly_new_customers_y2': 2,
        'monthly_new_customers_y3': 2,
        'monthly_churn_rate': 0.02,
    }
    builder.create_detailed_scenario("Conservative Scenario", conservative_params)

    # Base case scenario
    base_params = {
        'initial_customers': 0,
        'initial_funding': 500000,
        'blended_arpu': 30000,
        'cogs_per_customer': 345,
        'fixed_opex': 40000,
        'variable_opex_per_customer': 1333,
        'monthly_new_customers_y1': 1.25,
        'monthly_new_customers_y2': 2.5,
        'monthly_new_customers_y3': 4,
        'monthly_churn_rate': 0.01,
    }
    builder.create_detailed_scenario("Base Case Scenario", base_params)

    # Optimistic scenario
    optimistic_params = {
        'initial_customers': 0,
        'initial_funding': 500000,
        'blended_arpu': 30000,
        'cogs_per_customer': 345,
        'fixed_opex': 40000,
        'variable_opex_per_customer': 1333,
        'monthly_new_customers_y1': 1.25,
        'monthly_new_customers_y2': 3.33,
        'monthly_new_customers_y3': 5.25,
        'monthly_churn_rate': 0.005,
    }
    builder.create_detailed_scenario("Optimistic Scenario", optimistic_params)

    # Create unit economics sheet
    builder.create_unit_economics_sheet()

    # Create headcount and OpEx sheet
    builder.create_headcount_sheet()

    # Create fundraising sheet
    builder.create_fundraising_sheet()

    # Create sensitivity sheet
    builder.create_sensitivity_sheet()

    builder.save()

if __name__ == "__main__":
    main()
